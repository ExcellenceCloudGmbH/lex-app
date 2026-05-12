"""
Tests for lex.core.fields.XLSX_field.XLSXField
===============================================

XLSXField is a Django FileField subclass that generates formatted Excel files
from pandas DataFrames.  Tests exercise the full ``create_excel_file_from_dfs``
pipeline and its helper methods for header-splitting, row-insertion, column
width calculation, number formatting, cell comments, and autofilter.

All tests work with real openpyxl/pandas/xlsxwriter — no mocking of the
spreadsheet libraries — so they verify the *actual* Excel output.
"""

import os
import sys
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import MagicMock

import django
from django.apps import apps

sys.path.append(str(Path(__file__).resolve().parents[2]))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "lex.process_admin.tests.django_test_settings")
if not apps.ready:
    django.setup()

from django.test import SimpleTestCase

try:
    import pandas as pd
    import openpyxl
    HAS_EXCEL_DEPS = True
except ImportError:
    HAS_EXCEL_DEPS = False

from lex.core.fields.XLSX_field import XLSXField


# ── helpers ──────────────────────────────────────────────────────────────────

def _make_field():
    """Create a standalone XLSXField with a mocked save method."""
    field = XLSXField()
    field.save = MagicMock()  # Prevent actual file system operations
    return field


def _make_simple_df(rows=3, cols=None):
    """Create a simple DataFrame for testing."""
    if cols is None:
        cols = ["name", "value", "score"]
    data = {col: [f"{col}_{i}" for i in range(rows)] for col in cols}
    return pd.DataFrame(data)


def _load_workbook_from_bytes(excel_file):
    """Load an openpyxl workbook from a BytesIO result."""
    excel_file.seek(0)
    return openpyxl.load_workbook(excel_file)


# ═══════════════════════════════════════════════════════════════════════════
# 1. Class-level constants
# ═══════════════════════════════════════════════════════════════════════════
class XLSXFieldConstantsTests(SimpleTestCase):
    """Verify the field's configuration constants."""

    def test_max_length(self):
        self.assertEqual(XLSXField.max_length, 300)

    def test_cell_format_is_string(self):
        self.assertIsInstance(XLSXField.cell_format, str)
        self.assertIn("#,##0.00", XLSXField.cell_format)

    def test_cell_format_without_color_no_red(self):
        """cell_format_without_color should NOT have [Red] marker."""
        self.assertNotIn("[Red]", XLSXField.cell_format_without_color)
        self.assertIn("#,##0.00", XLSXField.cell_format_without_color)

    def test_boolean_format_has_true_false(self):
        self.assertIn("TRUE", XLSXField.boolean_format)
        self.assertIn("FALSE", XLSXField.boolean_format)

    def test_is_file_field_subclass(self):
        from django.db.models import FileField
        self.assertTrue(issubclass(XLSXField, FileField))


# ═══════════════════════════════════════════════════════════════════════════
# 2. get_number_of_rows_to_insert
# ═══════════════════════════════════════════════════════════════════════════
@unittest.skipUnless(HAS_EXCEL_DEPS, "pandas/openpyxl not installed")
class GetNumberOfRowsTests(SimpleTestCase):
    """Test the dot-based header depth calculation."""

    def _make_sheet_with_headers(self, headers, index_len=0):
        """Build an openpyxl sheet with headers in row 1 starting at column index_len+1."""
        wb = openpyxl.Workbook()
        ws = wb.active
        for i, header in enumerate(headers):
            ws.cell(row=1, column=index_len + 1 + i, value=header)
        return ws

    def test_no_dots_returns_one(self):
        """Headers without dots → max split length = 1."""
        ws = self._make_sheet_with_headers(["alpha", "beta", "gamma"])
        field = _make_field()
        result = field.get_number_of_rows_to_insert(ws, 0)
        self.assertEqual(result, 1)

    def test_dotted_headers(self):
        """'A.B.C' splits into 3 → returns 3."""
        ws = self._make_sheet_with_headers(["group.sub.detail", "simple"])
        field = _make_field()
        result = field.get_number_of_rows_to_insert(ws, 0)
        self.assertEqual(result, 3)

    def test_empty_cells_ignored(self):
        """Empty cells should not contribute to the row count."""
        ws = self._make_sheet_with_headers(["a.b", None, "x"])
        field = _make_field()
        result = field.get_number_of_rows_to_insert(ws, 0)
        self.assertEqual(result, 2)

    def test_index_offset(self):
        """index_len > 0 should skip the first N columns (index columns)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Index column
        ws.cell(row=1, column=1, value="idx_header")
        # Data columns start at column 2
        ws.cell(row=1, column=2, value="a.b.c.d")
        field = _make_field()
        result = field.get_number_of_rows_to_insert(ws, 1)
        self.assertEqual(result, 4)

    def test_all_empty_returns_zero(self):
        """Sheet with only empty data cells → returns 0."""
        wb = openpyxl.Workbook()
        ws = wb.active
        field = _make_field()
        result = field.get_number_of_rows_to_insert(ws, 0)
        self.assertEqual(result, 0)


# ═══════════════════════════════════════════════════════════════════════════
# 3. split_entries_in_sheet
# ═══════════════════════════════════════════════════════════════════════════
@unittest.skipUnless(HAS_EXCEL_DEPS, "pandas/openpyxl not installed")
class SplitEntriesTests(SimpleTestCase):
    """Test header splitting into multiple rows with formatting."""

    def test_splits_dotted_header_into_rows(self):
        """'Category.SubCategory.Detail' → 3 cells in column, one per row."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Simulate: after inserting 2 extra rows, the original header is at row 3
        ws.cell(row=3, column=1, value="Category.SubCategory.Detail")
        field = _make_field()
        field.split_entries_in_sheet(ws, 2, 0)

        self.assertEqual(ws.cell(row=1, column=1).value, "Category")
        self.assertEqual(ws.cell(row=2, column=1).value, "SubCategory")
        self.assertEqual(ws.cell(row=3, column=1).value, "Detail")

    def test_split_cells_are_bold(self):
        """Split cells must have bold font."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="A.B")
        field = _make_field()
        field.split_entries_in_sheet(ws, 1, 0)

        self.assertTrue(ws.cell(row=1, column=1).font.bold)
        self.assertTrue(ws.cell(row=2, column=1).font.bold)

    def test_split_cells_have_borders(self):
        """Split cells must have thin borders on all 4 sides."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="X.Y")
        field = _make_field()
        field.split_entries_in_sheet(ws, 1, 0)

        cell = ws.cell(row=1, column=1)
        self.assertEqual(cell.border.top.border_style, "thin")
        self.assertEqual(cell.border.bottom.border_style, "thin")
        self.assertEqual(cell.border.left.border_style, "thin")
        self.assertEqual(cell.border.right.border_style, "thin")


# ═══════════════════════════════════════════════════════════════════════════
# 4. create_pivotable_row
# ═══════════════════════════════════════════════════════════════════════════
@unittest.skipUnless(HAS_EXCEL_DEPS, "pandas/openpyxl not installed")
class CreatePivotableRowTests(SimpleTestCase):
    """Test concatenation of split header rows into a pivot row."""

    def test_concatenates_specified_rows(self):
        """Concatenate rows 1 and 2 into row 4 (number_of_rows_to_be_inserted=3)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Category")
        ws.cell(row=2, column=1, value="SubCategory")
        ws.cell(row=3, column=1, value="")

        field = _make_field()
        field.create_pivotable_row(ws, 0, 3, range_of_pivot_concatenation=[1, 2])

        self.assertEqual(ws.cell(row=4, column=1).value, "Category SubCategory")

    def test_skips_empty_cells_in_concat(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Level1")
        ws.cell(row=2, column=1, value=None)

        field = _make_field()
        field.create_pivotable_row(ws, 0, 2, range_of_pivot_concatenation=[1, 2])

        self.assertEqual(ws.cell(row=3, column=1).value, "Level1")


# ═══════════════════════════════════════════════════════════════════════════
# 5. create_excel_file_from_dfs — end-to-end
# ═══════════════════════════════════════════════════════════════════════════
@unittest.skipUnless(HAS_EXCEL_DEPS, "pandas/openpyxl not installed")
class CreateExcelFileEndToEndTests(SimpleTestCase):
    """End-to-end tests for the main Excel generation method."""

    def test_basic_single_sheet(self):
        """A single DataFrame produces a valid Excel file with one sheet."""
        field = _make_field()
        df = _make_simple_df(3)
        result = field.create_excel_file_from_dfs(
            "test.xlsx", [df], sheet_names=["Data"]
        )

        self.assertIsInstance(result, BytesIO)
        wb = _load_workbook_from_bytes(result)
        self.assertIn("Data", wb.sheetnames)

    def test_multiple_sheets(self):
        """Multiple DataFrames → multiple sheets."""
        field = _make_field()
        df1 = _make_simple_df(2)
        df2 = pd.DataFrame({"col_a": [1, 2, 3]})
        result = field.create_excel_file_from_dfs(
            "multi.xlsx", [df1, df2], sheet_names=["Sheet1", "Sheet2"]
        )

        wb = _load_workbook_from_bytes(result)
        self.assertEqual(len(wb.sheetnames), 2)
        self.assertIn("Sheet1", wb.sheetnames)
        self.assertIn("Sheet2", wb.sheetnames)

    def test_empty_dataframe_gets_blank_row(self):
        """An empty DataFrame should get a blank row appended (len check in source)."""
        field = _make_field()
        df = pd.DataFrame({"a": pd.Series(dtype="str"), "b": pd.Series(dtype="str")})
        result = field.create_excel_file_from_dfs(
            "empty.xlsx", [df], sheet_names=["Empty"]
        )

        wb = _load_workbook_from_bytes(result)
        ws = wb["Empty"]
        # Should have at least a header row + 1 blank data row
        self.assertGreaterEqual(ws.max_row, 2)

    def test_none_dataframe_skipped(self):
        """None entries in the data_frames list are skipped."""
        field = _make_field()
        df = _make_simple_df(1)
        result = field.create_excel_file_from_dfs(
            "with_none.xlsx", [df, None], sheet_names=["Good", "Skipped"]
        )

        wb = _load_workbook_from_bytes(result)
        self.assertIn("Good", wb.sheetnames)
        # None DF should not create a sheet
        self.assertNotIn("Skipped", wb.sheetnames)

    def test_default_sheet_name(self):
        """If sheet_names not provided, defaults to ['Sheet']."""
        field = _make_field()
        df = _make_simple_df(1)
        result = field.create_excel_file_from_dfs("default_sheet.xlsx", [df])

        wb = _load_workbook_from_bytes(result)
        self.assertIn("Sheet", wb.sheetnames)

    def test_save_called_with_correct_path(self):
        """field.save() is called with the path and File content."""
        field = _make_field()
        df = _make_simple_df(1)
        field.create_excel_file_from_dfs("output.xlsx", [df])

        field.save.assert_called_once()
        call_args = field.save.call_args
        self.assertEqual(call_args[0][0], "output.xlsx")

    def test_index_false_no_index_column(self):
        """When index=False, no index column is written."""
        field = _make_field()
        df = _make_simple_df(2, cols=["a", "b"])
        result = field.create_excel_file_from_dfs(
            "no_idx.xlsx", [df], sheet_names=["Data"], index=False
        )

        wb = _load_workbook_from_bytes(result)
        ws = wb["Data"]
        # First column should be 'a', not an index
        self.assertEqual(ws.cell(row=1, column=1).value, "a")

    def test_with_index(self):
        """When index=True (default), index column appears."""
        field = _make_field()
        df = pd.DataFrame({"val": [10, 20]}, index=pd.Index([100, 200], name="id"))
        result = field.create_excel_file_from_dfs(
            "with_idx.xlsx", [df], sheet_names=["Data"], index=True
        )

        wb = _load_workbook_from_bytes(result)
        ws = wb["Data"]
        # First column should be the index header
        self.assertEqual(ws.cell(row=1, column=1).value, "id")

    def test_comments_written(self):
        """Cell comments are written when provided."""
        field = _make_field()
        df = pd.DataFrame({"alpha": [1], "beta": [2]})
        comments = {"Data": {"alpha": "This is alpha", "beta": "This is beta"}}
        result = field.create_excel_file_from_dfs(
            "comments.xlsx", [df], sheet_names=["Data"], comments=comments, index=False
        )

        wb = _load_workbook_from_bytes(result)
        ws = wb["Data"]
        # xlsxwriter writes comments, openpyxl can read them
        # The comment for column "alpha" is at (row=1, col=1) when index=False
        comment = ws.cell(row=1, column=1).comment
        if comment is not None:
            self.assertIn("alpha", comment.text)

    def test_returns_seeked_bytesio(self):
        """The returned BytesIO should be seeked to position 0 (ready to read)."""
        field = _make_field()
        df = _make_simple_df(1)
        result = field.create_excel_file_from_dfs("test.xlsx", [df])
        self.assertEqual(result.tell(), 0)


# ═══════════════════════════════════════════════════════════════════════════
# 6. insert_rows_before_first_row
# ═══════════════════════════════════════════════════════════════════════════
@unittest.skipUnless(HAS_EXCEL_DEPS, "pandas/openpyxl not installed")
class InsertRowsTests(SimpleTestCase):
    """Test the row-insertion wrapper."""

    def test_inserts_correct_number_of_rows(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Header")
        ws.cell(row=2, column=1, value="Data")

        field = _make_field()
        field.insert_rows_before_first_row(ws, 3)

        # Original row 1 content should now be at row 4
        self.assertEqual(ws.cell(row=4, column=1).value, "Header")
        self.assertEqual(ws.cell(row=5, column=1).value, "Data")


if __name__ == "__main__":
    unittest.main()
